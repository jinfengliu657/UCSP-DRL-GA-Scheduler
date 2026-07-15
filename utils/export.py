"""
导出工具 - 将排课结果导出到Excel和数据库 (Fixed: 25 Time Slots)
文件名: utils/export.py
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict
from pathlib import Path
from datetime import datetime
import json

from models.chromosome import Chromosome
from data.data_loader import ScheduleData


class ScheduleExporter:
    """
    排课结果导出器
    Strictly enforces 5 days * 5 periods = 25 slots logic.
    """

    # --- 核心修正：严格定义的25个时间槽映射 ---
    TIME_SLOTS = {}
    DAYS = ["周一", "周二", "周三", "周四", "周五"]
    PERIODS = [1, 2, 3, 4, 5]

    # 自动生成 0-24 的映射表
    _idx = 0
    for day in DAYS:
        for p in PERIODS:
            TIME_SLOTS[_idx] = (day, p)
            _idx += 1

    def __init__(self, data: ScheduleData, output_dir: str = "results/exports"):
        self.data = data
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(self, chromosome: Chromosome, filename: str = None):
        """
        导出所有排课结果到一个Excel文件（包含多个Sheet）
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"schedule_result_{timestamp}.xlsx"

        filepath = self.output_dir / filename
        wb = Workbook()

        # 1. 详细排课列表 (Raw Data)
        self._create_detail_sheet(wb, chromosome)

        # 2. 班级课程表 (Matrix View)
        self._create_class_matrix_sheet(wb, chromosome)

        # 3. 教师课程表
        self._create_teacher_matrix_sheet(wb, chromosome)

        # 4. 统计信息
        self._create_stats_sheet(wb, chromosome)

        # 删除默认生成的空白Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filepath)
        print(f"排课结果已导出至 Excel: {filepath}")
        return filepath

    def _create_detail_sheet(self, wb, chromosome):
        """生成详细清单 Sheet"""
        ws = wb.create_sheet("详细排课列表")
        headers = ["任务ID", "课程名称", "班级", "教师", "教室", "人数", "周次", "节次", "时间槽ID"]
        ws.append(headers)

        for task_id, gene in chromosome.genes.items():
            if task_id not in self.data.tasks: continue

            task = self.data.tasks[task_id]
            slot = gene >> 16
            room_id = gene & 0xFFFF

            # --- 核心校验 ---
            if slot not in self.TIME_SLOTS:
                continue  # 跳过非法时间槽 (e.g. >24)

            day_str, period = self.TIME_SLOTS[slot]

            # 获取名称
            room_name = str(room_id)
            # 尝试从教室字典获取真实名称
            room_key = str(room_id)
            if room_key in self.data.classrooms:
                room_name = self.data.classrooms[room_key].classroom_name
            elif room_key.zfill(6) in self.data.classrooms:
                room_name = self.data.classrooms[room_key.zfill(6)].classroom_name

            row = [
                task.id,
                task.course_class_name,  # 或 task.course_id
                self.data.classes[task.class_id].class_name if task.class_id in self.data.classes else task.class_id,
                self.data.teachers[
                    task.teacher_id].teacher_name if task.teacher_id in self.data.teachers else task.teacher_id,
                room_name,
                task.student_count,
                day_str,
                f"第{period}节",
                slot
            ]
            ws.append(row)

    def _create_class_matrix_sheet(self, wb, chromosome):
        """生成班级课表矩阵"""
        ws = wb.create_sheet("班级课表矩阵")

        # 准备数据结构: class_id -> { (day, period): "Course(Room)" }
        schedules = {}

        for task_id, gene in chromosome.genes.items():
            if task_id not in self.data.tasks: continue
            task = self.data.tasks[task_id]

            slot = gene >> 16
            if slot not in self.TIME_SLOTS: continue

            day_str, period = self.TIME_SLOTS[slot]
            c_name = self.data.classes[
                task.class_id].class_name if task.class_id in self.data.classes else task.class_id

            if c_name not in schedules: schedules[c_name] = {}

            # 课程信息字符串
            info = f"{task.course_class_name}"
            schedules[c_name][(day_str, period)] = info

        # 写入Excel
        # 表头: 班级 | 周一1 | 周一2 ... | 周五5
        headers = ["班级"]
        time_keys = []
        for d in self.DAYS:
            for p in self.PERIODS:
                headers.append(f"{d}-{p}")
                time_keys.append((d, p))

        ws.append(headers)

        for c_name in sorted(schedules.keys()):
            row = [c_name]
            for tk in time_keys:
                row.append(schedules[c_name].get(tk, ""))
            ws.append(row)

    def _create_teacher_matrix_sheet(self, wb, chromosome):
        """生成教师课表矩阵"""
        ws = wb.create_sheet("教师课表矩阵")

        schedules = {}
        for task_id, gene in chromosome.genes.items():
            if task_id not in self.data.tasks: continue
            task = self.data.tasks[task_id]
            slot = gene >> 16
            if slot not in self.TIME_SLOTS: continue

            day_str, period = self.TIME_SLOTS[slot]
            t_name = self.data.teachers[
                task.teacher_id].teacher_name if task.teacher_id in self.data.teachers else task.teacher_id

            if t_name not in schedules: schedules[t_name] = {}
            schedules[t_name][(day_str, period)] = f"{task.course_class_name}"

        headers = ["教师"]
        time_keys = []
        for d in self.DAYS:
            for p in self.PERIODS:
                headers.append(f"{d}-{p}")
                time_keys.append((d, p))

        ws.append(headers)

        for t_name in sorted(schedules.keys()):
            row = [t_name]
            for tk in time_keys:
                row.append(schedules[t_name].get(tk, ""))
            ws.append(row)

    def _create_stats_sheet(self, wb, chromosome):
        """生成统计信息 Sheet"""
        ws = wb.create_sheet("统计面板")
        ws.append(["指标", "数值"])
        ws.append(["总适应度", chromosome.fitness])

        # 统计分布
        daily_counts = {d: 0 for d in self.DAYS}
        for gene in chromosome.genes.values():
            slot = gene >> 16
            if slot in self.TIME_SLOTS:
                day, _ = self.TIME_SLOTS[slot]
                daily_counts[day] += 1

        for day in self.DAYS:
            ws.append([f"{day}课程数", daily_counts[day]])

    def export_statistics(self, chromosome: Chromosome, output_path: str = None):
        """导出 JSON 统计"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.output_dir / f"statistics_{timestamp}.json"

        stats = {
            "fitness": float(chromosome.fitness),
            "total_tasks": len(chromosome.genes),
            "export_time": datetime.now().isoformat()
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(stats, f, ensure_ascii=False, indent=4)
        print(f"统计信息已导出: {output_path}")